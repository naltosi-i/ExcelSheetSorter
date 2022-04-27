'''エクセルのシートをソートリストに従って並べ替える。
並べ替えるとき、任意の位置以降に並べ替えたシートをまとめることができる。
元ファイルのコピーを作って並べ替えるのでデータ非破壊。
'''

import openpyxl
import os
import sys
import shutil
import pathlib
import random
sys.dont_write_bytecode = True # pycacheを作らないようにする

# 設定項目
path = 'dummy.xlsx' # 元ファイル
savename = "sorted_" + path # 完成ファイル
sortlist = 'BTSJsortList.xlsx' # ソートリストの入っているファイル
sl = 3 # 並べ替え先頭位置（0から始まるシートのインデックス。4枚目から始めたい場合は3）
# 設定ここまで

def makeSortList(sheet): # ソート用のリストをつくる。戻り値はインデックスのリストと、ソートキーのリストのタプル
    silist = [] #ソートリストのindexのリスト
    selist = [] #ソートリストの要素のリスト

    for row in sheet.iter_rows(min_row=2):
        # print(row[0].value)
        silist.append(row[0].value)
        selist.append(row[1].value)
    return silist, selist

def sheetSorter(Book, sheet, filename, sheetlocation):# シートの順番をソートリストに従い並べ替える。makeSortListを使う。
    wb = Book
    ss = sheet
    savename = filename
    sl = sheetlocation
    slist = makeSortList(ss)[1]
    print("シートを並べ替えています\n")

    for e in reversed(slist):
        
        if e in wb.sheetnames:
            print(e + " exists in sheets.")
            ws = wb[e]
            print("sheet index of " + e + " is " + str(wb.index(ws)))
            print(e + " is moving by offset=" + str( sl - wb.index(ws)))
            wb.move_sheet(ws, offset= sl - wb.index(ws))
            print("sheet index of " + e + " is " + str(wb.index(ws)) + "\n")
            wb.save(savename)
        else:
            continue
    print("完了")


try:
    shutil.copy(path, savename)
except PermissionError:
    print('PermissionError: Excelファイルを閉じてください。')
    exit()
else:
    pass

wb = openpyxl.load_workbook(savename)
slb = openpyxl.load_workbook(sortlist)


ss = slb.worksheets[0]
try:
    sheetSorter(wb, ss, savename, sl)
except PermissionError:
    print('PermissionError: Excelファイルを閉じてください。')
    exit()
else:
    pass